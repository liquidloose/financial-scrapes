from importlib.metadata import metadata
import re
from tabnanny import verbose
import numpy as np
import pandas as pd
import time
from pathlib import Path
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pyparsing import col


logging.basicConfig(level=logging.INFO, filename='app.log', filemode='a',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%d-%b-%y %H:%M:%S')

logger = logging.getLogger(__name__)


class ExcelSheetCreator:
    '''
    Takes in JSON, converts it into data frames and writes that
    data to an excel spreadsheet.
    '''

    def __init__(self, data):
        self.data = data
        # self.file_check()
        self.writer()

    def __str__(self):
        return "ExcelSheetCreator Object"

    @staticmethod
    def file_check():
        '''
        Checks if the excel file exists. If it doesn't exist, a header row
        is added to the spreadsheet.
        '''
        path_exists = Path("/var/www/financial-scrapes/test/wsj.xlsx")
        if path_exists.is_file():
            print('File exists!')
            logger.info('File exists!')
            return True
        else:
            logger.info('No file exists yet')
            print('No file exists')
            return False
            # self.writer()

    @staticmethod
    def wsj_data(data):
        '''
        Returns a two item list containing data related to NYSE and NASDAQ
        '''

        def drop(data_frame):
            data_frame.drop('previousClose', axis=1, inplace=True)
            data_frame.drop('weekAgo', axis=1, inplace=True)

        nyse_data = pd.DataFrame(data['nyse_data'])
        nasdaq_data = pd.DataFrame(data['nasdaq_data'])

        data = [nasdaq_data, nyse_data]
        for x in data:
            drop(x)

        # print(nyse_data)
        return [nyse_data.T, nasdaq_data.T]

    @staticmethod
    def data_injector(data_1, data_2):
        '''
        Inserts exchange names and scrape times into the data
        '''
        data_1.insert(
            0, " ", ["", "Exchange", "NYSE"])

        data_1.insert(
            0, " 2", ["", "Scraped On", time.strftime("%b %d, %Y")])

        data_1.insert(
            15, " ", ["", "Exchange", "NASDAQ"], allow_duplicates=True)

        data = pd.concat([data_1, data_2], axis=1)

        return data.iloc[1:3]

    @staticmethod
    def sheet_filter():
        '''
        Formats spreadsheet data types and column widths
        '''
        workbook = load_workbook('wsj.xlsx')
        sheet = workbook["stock_data"]

        def resize_columns():
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

        for row in sheet:
            for cell in row:
                cell_coordinate = cell.coordinate
                print(cell_coordinate)
                if isinstance(cell.value, str) and re.match(r"(\d+)(,)", cell.value):
                    strip_comma = int(cell.value.replace(",", ""))
                    sheet[cell_coordinate] = strip_comma
                elif isinstance(cell.value, str) and re.match(r"(\d+)(.)", cell.value):
                    cell_to_float = float(cell.value)
                    sheet[cell_coordinate] = cell_to_float
        resize_columns()
        workbook.save('wsj.xlsx')

    def writer(self):
        '''
        Writes the data to an excel file
        '''

        file_exists = self.file_check()
        print(f"does the file exist? {file_exists}")

        data = self.wsj_data(self.data)

        sheet_data = self.data_injector(data[0], data[1])

        if file_exists is False:
            print('Creating file and writing data to it')
            logger.info('Creating file and writing data to it')
            sheet_data.to_excel(
                'wsj.xlsx', sheet_name="stock_data", index=False, header=False)
        elif file_exists is True:
            print('Appending data to file')
            logger.info('Appending data to file')
            with pd.ExcelWriter('wsj.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                append_data = pd.DataFrame(sheet_data.iloc[1])
                transposed_appended_data = pd.DataFrame(sheet_data.iloc[1]).T
                print(f'this is the sheet data: {transposed_appended_data}')
                transposed_appended_data.to_excel(
                    writer, sheet_name="stock_data", startrow=writer.sheets["stock_data"].max_row,
                    index=False, header=False)

        self.sheet_filter()
